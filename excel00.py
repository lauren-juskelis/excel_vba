import openpyxl, os
os.chdir('C://Users//ljuskelis//Desktop//SIECA')

wb = openpyxl.load_workbook('CR - BZ EX.xlsx')
ns = wb.create_sheet('Sheet 1')

cs = wb['Sheet']
ns = wb['Sheet 1']

selection = cs['A3':'A4000']

def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected=[]
    for i in range(startRow, endRow + 1):
        rowSelected = []
        for j in range(startCol, endCol+1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        rangeSelected.append(rowSelected)

    return rangeSelected

def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow+1):
        countCol = 0
        for j in range(startCol,endCol+1):
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1



def createData():
    print("Processing...")
    selectedRange = copyRange(1,3,1,4000,cs)
    pastingRange = pasteRange(1,1,1,3998,ns,selectedRange)
    cs.save('CR - BZ EX.xlsx')
    print("Range copied and pasted!")

