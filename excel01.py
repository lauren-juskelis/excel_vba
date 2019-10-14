# Python script to extract only the first six characters from each cell in an entire column and 
# overwrite the previous contents in Excel. Loops through many files in a directory.

import openpyxl, os, re
import pandas as pd

def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected=[]
    for i in range(startRow, endRow + 1):
        rowSelected = []
        for j in range(startCol, endCol+1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        rangeSelected.append(rowSelected)
    return rangeSelected

def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow+1):
        countCol = 0
        for j in range(startCol,endCol+1):
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

def findFiles():
    files = []
    for file in os.listdir():
        if file.endswith('.xlsx'):
            files.append(file)
    return files
 
# Set directory.

os.chdir('C://Users//ljuskelis//Desktop//SIECA')

# Find file names in directory.

files = findFiles()

# Loop through files, perform tasks, and save.

for file in files:
    print('Working on file ' + file + '.')
    workbook = openpyxl.load_workbook(file)
    sheet = workbook['Sheet']
    data = copyRange(1,3,1,4000,sheet)
    df = pd.DataFrame(data, columns = ['long'])
    df['short'] = df['long'].astype(str).str[:6]
    df = df.replace('None', '')
    tempFile = file[:-5] + ' TEMP.xlsx'
    df.to_excel(tempFile)
    tempWorkbook = openpyxl.load_workbook(tempFile)
    tempSheet = tempWorkbook['Sheet1']
    digits = copyRange(3,1,3,3998, tempSheet)
    pasteRange(1, 3, 1, 4000, sheet, digits)
    workbook.save(file)
    os.remove(file[:-5] + ' TEMP.xlsx')

print('Finished.')

